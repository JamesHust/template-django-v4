import openpyxl
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError, NotFound, ParseError, PermissionDenied
from rest_framework.response import Response

from common.common import get_campaign_with_role, get_campaign_or_404
from common.permissions import AdminCampaignPermission, CampaignEditOrReadOnly
from voicebot_qc.api.user_campaign.serializers import UserCampaignSerializer, ImportPermissionsSerializer
from voicebot_qc.models import UserCampaign


class UserCampaignViewSet(viewsets.ModelViewSet):
    queryset = UserCampaign.objects.all()
    serializer_class = UserCampaignSerializer
    permission_classes = [AdminCampaignPermission]

    def get_queryset(self):
        return UserCampaign.objects.filter(campaign__id=self.kwargs['campaign_id_pk'], campaign__is_deleted=False)

    def perform_create(self, serializer):
        campaign = get_campaign_with_role(campaign_id=self.kwargs['campaign_id_pk'], user=self.request.user,
                                          method='POST', permission=UserCampaign.QC_LEAD)
        try:
            user = User.objects.get(email=serializer.validated_data.get('user').get('email'))
            if user.is_staff:
                raise ValidationError(detail='User is an admin')
        except ObjectDoesNotExist:
            raise NotFound(detail='Email does not exists')
        if user == campaign.created_by:
            raise ValidationError(detail='User is an owner')
        if UserCampaign.objects.filter(user=user, campaign=campaign).exists():
            raise ValidationError(detail='User Campaign already exists')
        serializer.save(
            campaign=campaign,
            user=user,
            created_by=self.request.user
        )


@api_view(['POST'])
@permission_classes([CampaignEditOrReadOnly])
def import_campaign_permission(request, campaign_id):
    serializer = ImportPermissionsSerializer(data=request.data)
    if not serializer.is_valid():
        raise ValidationError(detail=serializer.errors)
    file = serializer.validated_data.get('file')
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    if sheet['A1'].value == 'Email' and sheet['B1'].value == 'Permission':
        pass
    else:
        raise ValidationError(detail='Wrong template file')
    header = True
    import_fail = 0
    import_success = 0
    error = []
    for row in sheet.iter_rows():
        if header:
            header = False
            continue
        email = row[0].value
        role = row[1].value
        if role in [UserCampaign.QC_LEAD, UserCampaign.QC_MEMBER] and User.objects.filter(email__exact=email).exists():
            UserCampaign.objects.update_or_create(campaign_id=campaign_id, user=User.objects.get(email__exact=email),
                                                  created_by=request.user,
                                                  defaults={'permission': role})
            import_success += 1
        else:
            if role not in [UserCampaign.QC_LEAD, UserCampaign.QC_MEMBER]:
                error_msg = 'Permission not found'
            elif User.objects.filter(email__exact=email).exists() is False:
                error_msg = 'User Email not found'
            else:
                error_msg = 'Something wrong'
            import_fail += 1
            error.append({'line': row[0].row, 'msg': error_msg})

    return Response(data={'import_success': import_success, 'import_fail': {'count': import_fail, 'error': error}})


@api_view(['GET'])
@permission_classes([])
def template_import_campaign_permission(request, campaign_id):
    try:
        campaign = get_campaign_or_404(campaign_id=campaign_id, user=request.user)
    except ObjectDoesNotExist:
        raise NotFound()
    wb = Workbook(write_only=False)
    ws = wb.active
    header = ['Email', 'Permission']
    ws.append(header)
    ws1 = wb.create_sheet(title='type_of_permission')
    ws1['A1'] = 'qc_lead'
    ws1['A2'] = 'qc_member'
    response = HttpResponse(content=save_virtual_workbook(wb),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=import_permission_{}.xlsx'.format(slugify(campaign.name))
    return response


@api_view(['GET'])
def get_user_campaign_permission(request):
    pers = []
    for per in UserCampaign.PERMISSION_CHOICE:
        pers.append({
            "name": per[1],
            "permission": per[0]
        })
    return Response(pers)
