import json

from django.http import HttpResponse
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.decorators import api_view, permission_classes

from common.common import get_call_list_or_404
from voicebot_qc.models import DataInput, Call, Record


@api_view(['GET'])
@permission_classes([])
def export_report_batch(request, campaign_id, call_list_id):
    call_list = get_call_list_or_404(campaign_id=campaign_id, call_list_id=call_list_id)
    wb = Workbook()
    ws = wb.active
    ws.title = call_list.name
    ws.append(['Batch:', call_list.name])
    ws.append([])
    data_input_list = call_list.campaign.data_input.all()
    criteria_list = call_list.campaign.criterias.all().order_by('name')
    # customer_data = data_input_list.filter(source=DataInput.CUSTOMER_PROPERTIES)
    # system_data = data_input_list.filter(source=DataInput.SYSTEM)
    # data_collection_data = data_input_list.filter(source=DataInput.CUSTOMER_PROPERTIES)
    first_column = ['Call ID', 'Start Time']

    # title data input
    end_column_input = len(first_column) + len(data_input_list)
    ws['A3'] = 'DATA INPUT'
    ws['A3'].font = Font(size=12, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_column_input)

    # title data output
    start_column_output = 1 + end_column_input
    end_column_output = end_column_input + len(criteria_list)
    ws.cell(row=3, column=start_column_output).value = 'DATA OUTPUT'
    ws.cell(row=3, column=start_column_output).font = Font(size=12, bold=True)
    ws.cell(row=3, column=start_column_output).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=3, start_column=start_column_output, end_row=3, end_column=end_column_output)
    header = first_column + [data_input.field for data_input in data_input_list] + [criteria.name for criteria in
                                                                                    criteria_list]
    calls = call_list.calls.all()
    ws.append(header)

    # add data to sheet
    append_data_to_sheet(ws=ws, calls=calls, data_input_list=data_input_list, criteria_list=criteria_list)
    response = HttpResponse(content=save_virtual_workbook(wb),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report_batch_{}.xlsx'.format(slugify(call_list.name))
    return response


def append_data_to_sheet(ws, calls, data_input_list, criteria_list):
    for call in calls:
        record_list = []
        for criteria in criteria_list:
            try:
                record = Record.objects.get(criteria=criteria, call=call).result
            except:
                record = ''
            record_list.append(record)
        input_list = []
        for data_input in data_input_list:
            try:
                call_data = json.loads(call.meta_data)
            except:
                call_data = {}
            input_list.append(call_data.get(data_input.source).get(data_input.field))
        ws.append([call.call_id, str(call.start_time)] + input_list + record_list)
